#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o painel "Cobre Seu Politico" (HTML unico).

Fontes aceitas:
  python3 gerar_dashboard.py planilha.xlsx
  python3 gerar_dashboard.py --csv contatos.csv
  python3 gerar_dashboard.py --sheet "Cobre Seu Politico - Contatos"
      (usa a variavel de ambiente GOOGLE_CREDS com o JSON da conta de servico)
"""
import json, re, sys, os, csv, io, argparse, unicodedata
from datetime import datetime

COLUNAS = ["Casa", "Nome", "Partido", "UF", "E-mail oficial",
           "Telefone do gabinete", "Gabinete", "Endereço", "Página oficial"]


# ---------------------------------------------------------------- dados
def so_digitos(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\D", "", s)


def montar_registro(linhas):
    """linhas = lista de listas, primeira e o cabecalho."""
    if not linhas:
        return []
    cab = [str(c).strip() for c in linhas[0]]
    idx = {}
    for alvo in COLUNAS:
        for i, c in enumerate(cab):
            if strip_acc(c) == strip_acc(alvo):
                idx[alvo] = i
                break
    faltando = [c for c in COLUNAS if c not in idx and c != "Endereço"]
    if faltando:
        sys.exit("Colunas nao encontradas na planilha: " + ", ".join(faltando))

    def campo(linha, nome):
        i = idx.get(nome)
        if i is None or i >= len(linha):
            return ""
        v = linha[i]
        return "" if v is None else str(v).strip()

    reg = []
    for linha in linhas[1:]:
        if not linha or not campo(linha, "Nome"):
            continue
        casa = campo(linha, "Casa").lower()
        reg.append({
            "c": "S" if casa.startswith("sen") else "C",
            "n": campo(linha, "Nome"),
            "p": campo(linha, "Partido"),
            "u": campo(linha, "UF").upper(),
            "e": campo(linha, "E-mail oficial"),
            "t": so_digitos(campo(linha, "Telefone do gabinete")),
            "g": campo(linha, "Gabinete"),
            "l": campo(linha, "Página oficial"),
        })
    reg.sort(key=lambda x: (x["u"], x["c"], strip_acc(x["n"])))
    return reg


def de_xlsx(caminho):
    try:
        import openpyxl
    except ImportError:
        sys.exit("Instale openpyxl:  pip install openpyxl")
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["Contatos"] if "Contatos" in wb.sheetnames else wb[wb.sheetnames[0]]
    linhas = [list(r) for r in ws.iter_rows(values_only=True)]
    info = {}
    if "Info" in wb.sheetnames:
        for a, b in wb["Info"].iter_rows(values_only=True):
            if a:
                info[str(a).strip()] = str(b).strip() if b else ""
    return montar_registro(linhas), info


def de_csv(origem):
    if str(origem).startswith("http"):
        from urllib.request import urlopen
        texto = urlopen(origem, timeout=60).read().decode("utf-8")
    else:
        with open(origem, encoding="utf-8") as f:
            texto = f.read()
    linhas = list(csv.reader(io.StringIO(texto)))
    return montar_registro(linhas), {}


def de_sheets(nome_ou_chave):
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        sys.exit("Instale:  pip install gspread google-auth")
    bruto = os.environ.get("GOOGLE_CREDS")
    if not bruto:
        sys.exit("Defina a variavel de ambiente GOOGLE_CREDS com o JSON da conta de servico.")
    cred = Credentials.from_service_account_info(
        json.loads(bruto),
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly",
                "https://www.googleapis.com/auth/drive.readonly"])
    cli = gspread.authorize(cred)
    doc = (cli.open_by_key(nome_ou_chave)
           if len(nome_ou_chave) > 30 and " " not in nome_ou_chave
           else cli.open(nome_ou_chave))
    try:
        aba = doc.worksheet("Contatos")
    except Exception:
        aba = doc.get_worksheet(0)
    linhas = aba.get_all_values()
    info = {}
    try:
        for par in doc.worksheet("Info").get_all_values():
            if par and par[0]:
                info[par[0].strip()] = par[1].strip() if len(par) > 1 else ""
    except Exception:
        pass
    if not info.get("Última atualização"):
        info["Última atualização"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return montar_registro(linhas), info


def strip_acc(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s))
                   if unicodedata.category(c) != "Mn").lower()

# ---------------------------------------------------------------- html
TILEMAP = [
    ["", "", "RR", "AP", "", "", ""],
    ["", "AM", "PA", "MA", "CE", "RN", ""],
    ["AC", "RO", "TO", "PI", "PB", "PE", ""],
    ["", "MT", "GO", "DF", "BA", "AL", "SE"],
    ["", "MS", "MG", "ES", "", "", ""],
    ["", "", "SP", "RJ", "", "", ""],
    ["", "PR", "SC", "", "", "", ""],
    ["", "RS", "", "", "", "", ""],
]

UF_NOME = {
    "AC": "Acre", "AL": "Alagoas", "AP": "Amapá", "AM": "Amazonas", "BA": "Bahia",
    "CE": "Ceará", "DF": "Distrito Federal", "ES": "Espírito Santo", "GO": "Goiás",
    "MA": "Maranhão", "MT": "Mato Grosso", "MS": "Mato Grosso do Sul",
    "MG": "Minas Gerais", "PA": "Pará", "PB": "Paraíba", "PR": "Paraná",
    "PE": "Pernambuco", "PI": "Piauí", "RJ": "Rio de Janeiro",
    "RN": "Rio Grande do Norte", "RS": "Rio Grande do Sul", "RO": "Rondônia",
    "RR": "Roraima", "SC": "Santa Catarina", "SP": "São Paulo",
    "SE": "Sergipe", "TO": "Tocantins",
}

CSS = r"""
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --papel:#FBFBF9; --carta:#FFFFFF; --linha:#E5E4DE; --linha-forte:#D2D1C9;
  --tinta:#1B1C1E; --meio:#5F6166; --fraco:#8A8C91;
  --camara:#1F7A4C; --camara-luz:#EAF3EE;
  --senado:#1C5A9E; --senado-luz:#E9F0F8;
  --acao:#B0241A; --acao-luz:#FBEDEC;
  --r:9px; --sombra:0 1px 2px rgba(20,20,20,.05),0 1px 8px rgba(20,20,20,.04);
  --sans:"IBM Plex Sans",system-ui,-apple-system,sans-serif;
  --mono:"IBM Plex Mono",ui-monospace,"SF Mono",Menlo,monospace;
  --cond:"IBM Plex Sans Condensed","IBM Plex Sans",system-ui,sans-serif;
}
html{-webkit-text-size-adjust:100%}
body{font-family:var(--sans);background:var(--papel);color:var(--tinta);
  line-height:1.5;font-size:15px;padding:14px;
  -webkit-font-smoothing:antialiased;overflow-wrap:break-word}
.env{max-width:1080px;margin:0 auto}

/* ---------- cabecalho ---------- */
.topo{border-bottom:2px solid var(--tinta);padding-bottom:12px;margin-bottom:16px}
.selo{font-family:var(--mono);font-size:10.5px;letter-spacing:.16em;
  text-transform:uppercase;color:var(--acao);font-weight:600;margin-bottom:5px}
h1{font-family:var(--cond);font-size:clamp(27px,7.5vw,40px);font-weight:700;
  letter-spacing:-.015em;line-height:1.04}
.sub{color:var(--meio);font-size:13.5px;margin-top:6px;max-width:56ch}
.meta{display:flex;flex-wrap:wrap;gap:6px 14px;margin-top:11px;
  font-family:var(--mono);font-size:11px;color:var(--fraco)}
.meta b{color:var(--meio);font-weight:600}

/* ---------- mapa ---------- */
.bloco{margin-bottom:18px}
.rotulo{font-family:var(--mono);font-size:10.5px;letter-spacing:.13em;
  text-transform:uppercase;color:var(--fraco);font-weight:600;
  margin-bottom:9px;display:flex;align-items:center;gap:9px}
.rotulo::after{content:"";flex:1;height:1px;background:var(--linha)}
.mapa{display:grid;grid-template-columns:repeat(7,1fr);gap:5px;
  max-width:430px;margin-inline:auto}
.uf{aspect-ratio:1;border:1px solid var(--linha-forte);background:var(--carta);
  border-radius:6px;cursor:pointer;display:flex;flex-direction:column;
  align-items:center;justify-content:center;gap:0;font-family:var(--mono);
  transition:background .12s,border-color .12s,transform .12s;
  color:var(--tinta);padding:0;min-width:0}
.uf span{font-size:clamp(9.5px,2.6vw,12px);font-weight:600;line-height:1}
.uf small{font-size:clamp(8px,2.1vw,9.5px);color:var(--fraco);line-height:1.35;
  font-weight:400}
.uf:hover{border-color:var(--tinta);background:#F4F4F1}
.uf:focus-visible{outline:2px solid var(--acao);outline-offset:2px}
.uf[aria-pressed="true"]{background:var(--acao);border-color:var(--acao);color:#fff}
.uf[aria-pressed="true"] small{color:rgba(255,255,255,.75)}
.vazio-uf{visibility:hidden}

/* ---------- filtros ---------- */
.filtros{display:flex;flex-direction:column;gap:9px;margin-bottom:16px}
.busca{position:relative}
.busca input{width:100%;padding:11px 38px 11px 12px;border:1px solid var(--linha-forte);
  border-radius:var(--r);font:inherit;font-size:15px;background:var(--carta);
  color:var(--tinta)}
.busca input:focus{outline:2px solid var(--acao);outline-offset:-1px;border-color:var(--acao)}
.busca button{position:absolute;right:5px;top:50%;transform:translateY(-50%);
  border:0;background:none;cursor:pointer;color:var(--fraco);font-size:19px;
  width:30px;height:30px;line-height:1;border-radius:5px}
.busca button:hover{color:var(--acao);background:var(--acao-luz)}
.linha-f{display:flex;gap:9px;flex-wrap:wrap}
.seg{display:flex;border:1px solid var(--linha-forte);border-radius:var(--r);
  overflow:hidden;background:var(--carta);flex:1;min-width:210px}
.seg button{flex:1;border:0;background:none;padding:9px 6px;cursor:pointer;
  font:inherit;font-size:13px;font-weight:500;color:var(--meio);
  border-right:1px solid var(--linha);transition:background .12s,color .12s}
.seg button:last-child{border-right:0}
.seg button:hover{background:#F4F4F1}
.seg button[aria-pressed="true"]{background:var(--tinta);color:#fff}
.seg button[data-c="C"][aria-pressed="true"]{background:var(--camara)}
.seg button[data-c="S"][aria-pressed="true"]{background:var(--senado)}
select{padding:9px 10px;border:1px solid var(--linha-forte);border-radius:var(--r);
  font:inherit;font-size:13.5px;background:var(--carta);color:var(--tinta);
  cursor:pointer;flex:1;min-width:140px}
select:focus{outline:2px solid var(--acao);outline-offset:-1px}

/* ---------- barra de resultado ---------- */
.barra{display:flex;align-items:center;justify-content:space-between;gap:10px;
  flex-wrap:wrap;padding-bottom:9px;border-bottom:1px solid var(--linha);
  margin-bottom:13px}
.conta{font-family:var(--mono);font-size:12px;color:var(--meio)}
.conta b{color:var(--tinta);font-size:15px;font-weight:600}
.limpar{border:1px solid var(--linha-forte);background:var(--carta);
  border-radius:6px;padding:6px 11px;cursor:pointer;font:inherit;font-size:12px;
  color:var(--meio)}
.limpar:hover{border-color:var(--acao);color:var(--acao);background:var(--acao-luz)}

/* ---------- cartoes ---------- */
.lista{display:grid;gap:9px}
@media(min-width:660px){.lista{grid-template-columns:repeat(2,1fr)}}
@media(min-width:1000px){.lista{grid-template-columns:repeat(3,1fr)}}
.card{background:var(--carta);border:1px solid var(--linha);border-radius:var(--r);
  padding:12px 13px;box-shadow:var(--sombra);border-left:3px solid var(--camara);
  display:flex;flex-direction:column;gap:9px}
.card.s{border-left-color:var(--senado)}
.nome{font-weight:600;font-size:15.5px;line-height:1.25;letter-spacing:-.01em}
.chips{display:flex;flex-wrap:wrap;gap:5px;margin-top:5px}
.chip{font-family:var(--mono);font-size:10px;font-weight:500;padding:2.5px 6px;
  border-radius:4px;background:#F2F2EF;color:var(--meio);
  letter-spacing:.03em;white-space:nowrap}
.chip.casa{background:var(--camara-luz);color:var(--camara);font-weight:600}
.card.s .chip.casa{background:var(--senado-luz);color:var(--senado)}
.gab{font-family:var(--mono);font-size:11px;color:var(--fraco);line-height:1.45}
.acoes{display:flex;gap:6px;margin-top:auto;flex-wrap:wrap}
.b{flex:1;min-width:0;display:inline-flex;align-items:center;justify-content:center;
  gap:5px;padding:9px 8px;border-radius:6px;border:1px solid var(--linha-forte);
  background:var(--carta);color:var(--tinta);text-decoration:none;cursor:pointer;
  font:inherit;font-size:12.5px;font-weight:500;min-height:38px;
  transition:background .12s,border-color .12s,color .12s;white-space:nowrap}
.b:hover{border-color:var(--acao);color:var(--acao);background:var(--acao-luz)}
.b:focus-visible{outline:2px solid var(--acao);outline-offset:1px}
.b svg{width:14px;height:14px;flex:none}
.b.ok{background:var(--camara-luz);border-color:var(--camara);color:var(--camara)}
.b.mudo{color:var(--fraco);cursor:default;opacity:.55}
.b.mudo:hover{border-color:var(--linha-forte);color:var(--fraco);background:var(--carta)}

/* ---------- estados ---------- */
.convite{text-align:center;padding:34px 18px;border:1px dashed var(--linha-forte);
  border-radius:var(--r);background:var(--carta);color:var(--meio)}
.convite strong{display:block;color:var(--tinta);font-size:16px;
  margin-bottom:5px;font-weight:600}
.convite p{font-size:13.5px;max-width:40ch;margin-inline:auto}
.mais{width:100%;padding:12px;margin-top:11px;border:1px solid var(--linha-forte);
  background:var(--carta);border-radius:var(--r);cursor:pointer;font:inherit;
  font-size:13.5px;font-weight:500;color:var(--meio)}
.mais:hover{border-color:var(--tinta);color:var(--tinta);background:#F4F4F1}
.rodape{margin-top:22px;padding-top:13px;border-top:1px solid var(--linha);
  font-size:11.5px;color:var(--fraco);line-height:1.6}
.rodape a{color:var(--meio)}
.aviso{position:fixed;left:50%;bottom:16px;transform:translateX(-50%) translateY(90px);
  background:var(--tinta);color:#fff;padding:9px 16px;border-radius:99px;
  font-size:13px;font-weight:500;transition:transform .22s;z-index:9;
  box-shadow:0 4px 14px rgba(0,0,0,.2);pointer-events:none}
.aviso.on{transform:translateX(-50%) translateY(0)}
@media(prefers-reduced-motion:reduce){*{transition:none!important;animation:none!important}}
"""

JS = r"""
const $=s=>document.querySelector(s);
let uf="",casa="",part="",busca="",limite=48;
const sa=s=>s.normalize("NFD").replace(/[\u0300-\u036f]/g,"").toLowerCase();
const tel=t=>t&&t.length===8?"(61) "+t.slice(0,4)+"-"+t.slice(4):(t||"");
const telLink=t=>t&&t.length===8?"+5561"+t:"";
const esc=s=>(s||"").replace(/[&<>"]/g,c=>({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]));

const ICO={
 mail:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="2" y="4" width="20" height="16" rx="2"/><path d="m2 7 10 6 10-6"/></svg>',
 fone:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M22 16.9v3a2 2 0 0 1-2.2 2 19.8 19.8 0 0 1-8.6-3.1 19.5 19.5 0 0 1-6-6A19.8 19.8 0 0 1 2.1 4.2 2 2 0 0 1 4.1 2h3a2 2 0 0 1 2 1.7c.1 1 .4 1.9.7 2.8a2 2 0 0 1-.5 2.1L8.1 9.9a16 16 0 0 0 6 6l1.3-1.3a2 2 0 0 1 2.1-.4c.9.3 1.8.6 2.8.7a2 2 0 0 1 1.7 2z"/></svg>',
 link:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M18 13v6a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h6"/><path d="M15 3h6v6"/><path d="M10 14 21 3"/></svg>',
 copy:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="9" y="9" width="13" height="13" rx="2"/><path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"/></svg>'
};

function filtrar(){
  const b=sa(busca.trim());
  return DADOS.filter(d=>
    (!uf||d.u===uf)&&(!casa||d.c===casa)&&(!part||d.p===part)&&
    (!b||sa(d.n).includes(b)||sa(d.p).includes(b)));
}

function cartao(d){
  const casaN=d.c==="S"?"Senado":"Câmara";
  const t=tel(d.t),tl=telLink(d.t);
  return `<article class="card ${d.c==="S"?"s":""}">
   <div><h3 class="nome">${esc(d.n)}</h3>
    <div class="chips"><span class="chip casa">${casaN}</span>
     <span class="chip">${esc(d.p)}</span><span class="chip">${esc(d.u)}</span></div></div>
   ${d.g?`<div class="gab">${esc(d.g)}</div>`:""}
   <div class="acoes">
    ${d.e?`<a class="b" href="mailto:${esc(d.e)}">${ICO.mail}E-mail</a>
      <button class="b" data-copia="${esc(d.e)}" title="Copiar ${esc(d.e)}">${ICO.copy}Copiar</button>`:""}
    ${tl?`<a class="b" href="tel:${tl}">${ICO.fone}${t}</a>`:`<span class="b mudo">${ICO.fone}sem telefone</span>`}
    ${d.l?`<a class="b" href="${esc(d.l)}" target="_blank" rel="noopener">${ICO.link}Perfil</a>`:""}
   </div></article>`;
}

function pintar(){
  const r=filtrar(),alvo=$("#lista"),barra=$("#barra");
  document.querySelectorAll(".uf").forEach(b=>b.setAttribute("aria-pressed",b.dataset.uf===uf));
  document.querySelectorAll(".seg button").forEach(b=>b.setAttribute("aria-pressed",(b.dataset.c||"")===casa));
  const ativo=uf||casa||part||busca.trim();
  barra.style.display=ativo?"flex":"none";
  if(!ativo){
    alvo.innerHTML=`<div class="convite"><strong>Comece pelo seu estado</strong>
      <p>Toque na sigla do seu estado no mapa acima para ver quem representa você —
      ou busque pelo nome. São ${DADOS.length} parlamentares com contato oficial.</p></div>`;
    $("#mais").style.display="none";return;
  }
  let rot=[];
  if(uf)rot.push(UFN[uf]||uf);
  if(casa)rot.push(casa==="S"?"Senado":"Câmara");
  if(part)rot.push(part);
  $("#conta").innerHTML=`<b>${r.length}</b> ${r.length===1?"parlamentar":"parlamentares"}${rot.length?" · "+rot.join(" · "):""}`;
  if(!r.length){
    alvo.innerHTML=`<div class="convite"><strong>Nada encontrado</strong>
      <p>Nenhum parlamentar corresponde a esses filtros. Tente limpar algum deles.</p></div>`;
    $("#mais").style.display="none";return;
  }
  alvo.innerHTML=r.slice(0,limite).map(cartao).join("");
  const resta=r.length-limite;
  $("#mais").style.display=resta>0?"block":"none";
  if(resta>0)$("#mais").textContent=`Mostrar mais ${Math.min(resta,48)} de ${resta}`;
}

function aviso(t){const a=$("#aviso");a.textContent=t;a.classList.add("on");
  clearTimeout(a._t);a._t=setTimeout(()=>a.classList.remove("on"),1700);}

function viaCampo(txt){                       // usado quando o iframe bloqueia a API
  return new Promise((ok,falha)=>{
    const a=document.createElement("textarea");
    a.value=txt;a.setAttribute("readonly","");
    a.style.cssText="position:fixed;top:-9999px;opacity:0";
    document.body.appendChild(a);a.select();a.setSelectionRange(0,99999);
    let deu=false;
    try{deu=document.execCommand("copy")}catch(e){}
    document.body.removeChild(a);
    deu?ok():falha();
  });
}

function copiar(txt){
  if(navigator.clipboard&&window.isSecureContext)
    return navigator.clipboard.writeText(txt).catch(()=>viaCampo(txt));
  return viaCampo(txt);
}

function verResultados(){                     // no celular, leva direto à lista
  if(window.innerWidth>=660)return;
  const alvo=document.querySelector("#barra");
  if(alvo&&alvo.style.display!=="none")
    alvo.scrollIntoView({behavior:"smooth",block:"start"});
}

document.addEventListener("click",e=>{
  const u=e.target.closest(".uf");
  if(u){uf=uf===u.dataset.uf?"":u.dataset.uf;limite=48;pintar();
    if(uf)verResultados();return;}
  const s=e.target.closest(".seg button");
  if(s){casa=(s.dataset.c||"")===casa?"":(s.dataset.c||"");limite=48;pintar();return;}
  const c=e.target.closest("[data-copia]");
  if(c){copiar(c.dataset.copia).then(()=>{
    c.classList.add("ok");aviso("E-mail copiado");
    setTimeout(()=>c.classList.remove("ok"),1200);
  }).catch(()=>aviso("Copie manualmente: "+c.dataset.copia));}
});
$("#q").addEventListener("input",e=>{busca=e.target.value;limite=48;pintar()});
$("#limpaq").addEventListener("click",()=>{busca="";$("#q").value="";pintar();$("#q").focus()});
$("#partido").addEventListener("change",e=>{part=e.target.value;limite=48;pintar()});
$("#mais").addEventListener("click",()=>{limite+=48;pintar()});
$("#limpar").addEventListener("click",()=>{uf=casa=part=busca="";$("#q").value="";
  $("#partido").value="";limite=48;pintar();window.scrollTo({top:0,behavior:"smooth"})});
pintar();
"""


def montar(reg, info, saida):
    por_uf = {}
    for d in reg:
        por_uf[d["u"]] = por_uf.get(d["u"], 0) + 1
    partidos = sorted({d["p"] for d in reg})
    n_cam = sum(1 for d in reg if d["c"] == "C")
    n_sen = len(reg) - n_cam

    tiles = []
    for linha in TILEMAP:
        for uf in linha:
            if not uf:
                tiles.append('<span class="uf vazio-uf" aria-hidden="true"></span>')
            else:
                tiles.append(
                    f'<button class="uf" data-uf="{uf}" aria-pressed="false" '
                    f'title="{UF_NOME.get(uf, uf)}"><span>{uf}</span>'
                    f'<small>{por_uf.get(uf, 0)}</small></button>'
                )
    mapa = "\n".join(tiles)
    opts = "\n".join(f'<option value="{p}">{p}</option>' for p in partidos)

    atual = info.get("Última atualização", "")
    try:
        atual = datetime.fromisoformat(atual).strftime("%d/%m/%Y")
    except Exception:
        atual = (atual or "").split(" ")[0]

    html = f"""<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>Cobre Seu Político — contatos oficiais do Congresso</title>
<meta name="description" content="Busque deputados federais e senadores por estado e partido, com e-mail e telefone oficiais.">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans+Condensed:wght@600;700&family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body>
<div class="env">

<header class="topo">
  <div class="selo">Congresso Nacional · contatos oficiais</div>
  <h1>Cobre seu político</h1>
  <p class="sub">Quem ocupa um mandato responde a quem o elegeu. Encontre os
    parlamentares do seu estado e fale com o gabinete.</p>
  <div class="meta">
    <span><b>{len(reg)}</b> parlamentares</span>
    <span><b>{n_cam}</b> deputados</span>
    <span><b>{n_sen}</b> senadores</span>
    {f'<span>atualizado em <b>{atual}</b></span>' if atual else ''}
  </div>
</header>

<section class="bloco">
  <div class="rotulo">Escolha o estado</div>
  <div class="mapa" role="group" aria-label="Estados do Brasil">
{mapa}
  </div>
</section>

<section class="filtros">
  <div class="busca">
    <input id="q" type="search" placeholder="Buscar por nome ou partido…"
      aria-label="Buscar parlamentar por nome ou partido" autocomplete="off">
    <button id="limpaq" type="button" aria-label="Limpar busca">&times;</button>
  </div>
  <div class="linha-f">
    <div class="seg" role="group" aria-label="Filtrar por casa legislativa">
      <button type="button" data-c="" aria-pressed="true">Todos</button>
      <button type="button" data-c="C" aria-pressed="false">Câmara</button>
      <button type="button" data-c="S" aria-pressed="false">Senado</button>
    </div>
    <select id="partido" aria-label="Filtrar por partido">
      <option value="">Todos os partidos</option>
{opts}
    </select>
  </div>
</section>

<div class="barra" id="barra" style="display:none">
  <div class="conta" id="conta"></div>
  <button class="limpar" id="limpar" type="button">Limpar filtros</button>
</div>

<main class="lista" id="lista"></main>
<button class="mais" id="mais" type="button" style="display:none">Mostrar mais</button>

<footer class="rodape">
  <strong>Endereços:</strong> Câmara dos Deputados e Senado Federal —
  Praça dos Três Poderes, Brasília/DF.<br>
  Dados públicos da Câmara dos Deputados e do Senado Federal. Apenas contatos
  institucionais de gabinete. Ligações para Brasília usam o DDD 61.
</footer>
</div>

<div class="aviso" id="aviso" role="status" aria-live="polite"></div>

<script>
const DADOS={json.dumps(reg, ensure_ascii=False, separators=(',', ':'))};
const UFN={json.dumps(UF_NOME, ensure_ascii=False, separators=(',', ':'))};
{JS}
</script>
</body>
</html>"""
    with open(saida, "w", encoding="utf-8") as f:
        f.write(html)
    return len(html)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Gera o painel Cobre Seu Politico.")
    ap.add_argument("planilha", nargs="?", help="arquivo .xlsx")
    ap.add_argument("--csv", help="arquivo .csv ou URL de CSV publicado")
    ap.add_argument("--sheet", help="nome ou ID da planilha no Google Sheets")
    ap.add_argument("-o", "--saida", default="index.html")
    a = ap.parse_args()

    if a.sheet:
        reg, info = de_sheets(a.sheet)
    elif a.csv:
        reg, info = de_csv(a.csv)
    elif a.planilha:
        reg, info = de_xlsx(a.planilha)
    else:
        ap.error("informe um .xlsx, --csv ou --sheet")

    if not reg:
        sys.exit("Nenhum parlamentar encontrado na fonte de dados.")
    n = montar(reg, info, a.saida)
    print(f"{len(reg)} parlamentares -> {a.saida} ({n/1024:.0f} KB)")
